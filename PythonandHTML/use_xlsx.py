import openpyxl as exl

# 写入xlsx
wb = exl.Workbook()  # 定义工作表
sht = wb.active  # 获取活动表
sht.title = '1st'  # 给活动（工作）表重命名
sht['A1'] = '0'  # 赋值“A1”单元格（一（A）列1行）
sht['A1'] = '1'  # 同上，不过会覆盖原内容
w = [['2', '3'], ['4', '5']]  # 将写入内容放入变量
for i in w:
    sht.append(i)  # 将内容逐个读取（单层表格）并逐行写入
wb.save('1.xlsx')  # 保存

# 读取xlsx
wb = exl.load_workbook('1.xlsx')  # 打开目标工作簿
shtn = wb.sheetnames  # 查询所有工作表的名称
print(shtn)  # 输出查询结果
sht = wb['1st']  # 打开目标工作表
cl = sht['A1'].value  # 取出目标单元格并取值
print(cl)  # 输出值